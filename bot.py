# -*- coding: utf-8 -*-
import os
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
from openpyxl import load_workbook

# ===== CONFIG =====
BOT_TOKEN = "YOUR_BOT_TOKEN_HERE"
ADMIN_ID = 7867235273  # Admin এর Telegram ID

# ===== VARIABLES =====
ADMIN_FILE = None  # Admin আপলোড করা ফাইল

# ===== KEYBOARD =====
main_keyboard = ReplyKeyboardMarkup(
    [
        [KeyboardButton("Clear Admin File")],
    ],
    resize_keyboard=True
)

# ===== START COMMAND =====
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Welcome! Send Your File Here",
        reply_markup=main_keyboard
    )

# ===== Admin File Upload =====
async def upload_admin_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global ADMIN_FILE
    if update.message.from_user.id != ADMIN_ID:
        await update.message.reply_text("You are not authorized to upload admin file.")
        return

    if not update.message.document:
        await update.message.reply_text("Please send an Excel file.")
        return

    file = await update.message.document.get_file()
    file_path = f"admin_file.xlsx"
    await file.download_to_drive(file_path)
    ADMIN_FILE = file_path
    await update.message.reply_text("Admin file uploaded successfully.")

# ===== User File Check =====
async def check_user_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message.document:
        return  # যদি user file না দেয়, কিছু হবে না

    if not ADMIN_FILE:
        await update.message.reply_text("Admin file is not set yet.")
        return

    # Download user file
    file = await update.message.document.get_file()
    user_file_path = f"user_{update.message.from_user.id}.xlsx"
    await file.download_to_drive(user_file_path)

    # Load admin file (Column A only, ignore empty/None)
    wb_admin = load_workbook(ADMIN_FILE)
    ws_admin = wb_admin.active
    admin_numbers = {
        str(ws_admin[f"A{i}"].value).strip()
        for i in range(1, ws_admin.max_row + 1)
        if ws_admin[f"A{i}"].value not in (None, "")
    }

    # Load user file (Column A only, ignore empty/None)
    wb_user = load_workbook(user_file_path)
    ws_user = wb_user.active
    user_numbers = {
        str(ws_user[f"A{i}"].value).strip()
        for i in range(1, ws_user.max_row + 1)
        if ws_user[f"A{i}"].value not in (None, "")
    }

    # Check matching numbers
    matched_numbers = admin_numbers & user_numbers
    matched_count = len(matched_numbers)
    unmatched_count = len(user_numbers - admin_numbers)

    # Send formatted report
    await update.message.reply_text(
        f"YOUR REPORT: {matched_count} OK✅\nBACK 👉 : {unmatched_count} BACK❌"
    )

    # Notify admin
    await context.bot.send_message(
        chat_id=ADMIN_ID,
        text=f"User {update.message.from_user.id} checked a file.\nMatching numbers: {matched_count}"
    )

# ===== Clear Admin File =====
async def clear_admin_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global ADMIN_FILE
    if update.message.from_user.id != ADMIN_ID:
        await update.message.reply_text("You are not authorized to clear admin file.")
        return

    ADMIN_FILE = None
    await update.message.reply_text("Admin file cleared successfully.")

# ===== Main =====
if __name__ == "__main__":
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    # Handlers
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.Document.ALL & filters.User(ADMIN_ID), upload_admin_file))
    app.add_handler(MessageHandler(filters.Document.ALL & ~filters.User(ADMIN_ID), check_user_file))
    app.add_handler(MessageHandler(filters.Regex("Clear Admin File"), clear_admin_file))

    print("Bot is running...")
    app.run_polling()